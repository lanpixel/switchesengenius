import openpyxl
import csv

# Leer el archivo Excel
excel_file = 'EnGenius Switch Specs Comparison.xlsx'
wb = openpyxl.load_workbook(excel_file)

print(f"Hojas disponibles: {wb.sheetnames}")
print(f"Total de hojas: {len(wb.sheetnames)}\n")

# Convertir cada hoja a un CSV separado
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Crear nombre de archivo CSV seguro
    safe_name = sheet_name.replace('/', '_').replace('\\', '_')
    csv_file = f'EnGenius_Switch_Specs_{safe_name}.csv'

    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    print(f"Convertida hoja '{sheet_name}' a {csv_file}")
    print(f"  Filas: {ws.max_row}, Columnas: {ws.max_column}")

print("\nConversión completada!")
